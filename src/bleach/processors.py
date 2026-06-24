from __future__ import annotations

import tempfile
from collections import Counter
from pathlib import Path

import fitz
from openpyxl import load_workbook

from bleach.detectors import Span, detect_text
from bleach.learned import LearnedValue
from bleach.masking import mask_value


TEXT_EXTENSIONS = {".txt", ".log", ".md", ".csv", ".tsv"}
XLSX_EXTENSIONS = {".xlsx"}
PDF_EXTENSIONS = {".pdf"}
SUPPORTED_EXTENSIONS = TEXT_EXTENSIONS | XLSX_EXTENSIONS | PDF_EXTENSIONS


def process_file(
    source: Path,
    dest: Path,
    *,
    profile: str,
    learned_values: list[LearnedValue],
) -> dict[str, int]:
    if source.suffix.lower() in TEXT_EXTENSIONS:
        return _process_text_file(source, dest, profile=profile, learned_values=learned_values)
    if source.suffix.lower() in XLSX_EXTENSIONS:
        return _process_xlsx_file(source, dest, profile=profile, learned_values=learned_values)
    if source.suffix.lower() in PDF_EXTENSIONS:
        return _process_pdf_file(source, dest, profile=profile, learned_values=learned_values)
    raise ValueError("unsupported file type")


def verify_file(
    path: Path,
    *,
    profile: str,
    learned_values: list[LearnedValue],
) -> dict[str, int]:
    if path.suffix.lower() in TEXT_EXTENSIONS:
        text = path.read_text(encoding="utf-8")
        spans = detect_text(text, profile=profile, learned_values=learned_values)
        return dict(Counter(span.kind for span in spans))
    if path.suffix.lower() in XLSX_EXTENSIONS:
        return _verify_xlsx_file(path, profile=profile, learned_values=learned_values)
    if path.suffix.lower() in PDF_EXTENSIONS:
        return _verify_pdf_file(path, profile=profile, learned_values=learned_values)
    raise ValueError("unsupported file type")


def redact_text(
    text: str,
    *,
    profile: str,
    learned_values: list[LearnedValue],
) -> tuple[str, dict[str, int]]:
    spans = detect_text(text, profile=profile, learned_values=learned_values)
    redacted = _replace_spans(text, spans, profile)
    return redacted, dict(Counter(span.kind for span in spans))


def _process_text_file(
    source: Path,
    dest: Path,
    *,
    profile: str,
    learned_values: list[LearnedValue],
) -> dict[str, int]:
    text = source.read_text(encoding="utf-8")
    redacted, counts = redact_text(text, profile=profile, learned_values=learned_values)
    _atomic_write(dest, redacted)
    return counts


def _process_xlsx_file(
    source: Path,
    dest: Path,
    *,
    profile: str,
    learned_values: list[LearnedValue],
) -> dict[str, int]:
    workbook = load_workbook(source)
    counts: Counter[str] = Counter()
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    redacted, detected = redact_text(
                        cell.value,
                        profile=profile,
                        learned_values=learned_values,
                    )
                    if detected:
                        cell.value = redacted
                        counts.update(detected)
    _atomic_save_workbook(dest, workbook)
    return dict(counts)


def _verify_xlsx_file(
    path: Path,
    *,
    profile: str,
    learned_values: list[LearnedValue],
) -> dict[str, int]:
    workbook = load_workbook(path, data_only=False)
    counts: Counter[str] = Counter()
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    spans = detect_text(cell.value, profile=profile, learned_values=learned_values)
                    counts.update(span.kind for span in spans)
    return dict(counts)


def _process_pdf_file(
    source: Path,
    dest: Path,
    *,
    profile: str,
    learned_values: list[LearnedValue],
) -> dict[str, int]:
    counts: Counter[str] = Counter()
    with fitz.open(source) as doc:
        _reject_unsafe_pdf(doc)
        saw_extractable_text = False
        for page in doc:
            text = page.get_text()
            if text.strip():
                saw_extractable_text = True
            spans = detect_text(text, profile=profile, learned_values=learned_values)
            counts.update(span.kind for span in spans)
            for span in spans:
                for rect in page.search_for(span.text):
                    page.add_redact_annot(
                        rect,
                        text=mask_value(span.text, span.kind, profile),
                        fill=(1, 1, 1),
                        text_color=(0, 0, 0),
                    )
            if spans:
                page.apply_redactions()
        if not saw_extractable_text:
            raise ValueError("unsupported PDF without extractable text")
        doc.set_metadata({})
        _atomic_save_pdf(dest, doc)
    return dict(counts)


def _verify_pdf_file(
    path: Path,
    *,
    profile: str,
    learned_values: list[LearnedValue],
) -> dict[str, int]:
    counts: Counter[str] = Counter()
    with fitz.open(path) as doc:
        _reject_unsafe_pdf(doc)
        saw_extractable_text = False
        for page in doc:
            text = page.get_text()
            if text.strip():
                saw_extractable_text = True
            spans = detect_text(text, profile=profile, learned_values=learned_values)
            counts.update(span.kind for span in spans)
        if not saw_extractable_text:
            raise ValueError("unsupported PDF without extractable text")
    return dict(counts)


def _replace_spans(text: str, spans: list[Span], profile: str) -> str:
    result = text
    for span in sorted(spans, key=lambda item: item.start, reverse=True):
        result = (
            result[: span.start]
            + mask_value(span.text, span.kind, profile)
            + result[span.end :]
        )
    return result


def _atomic_write(dest: Path, text: str) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        "w",
        encoding="utf-8",
        dir=dest.parent,
        delete=False,
    ) as handle:
        temp_path = Path(handle.name)
        handle.write(text)
    temp_path.replace(dest)


def _atomic_save_workbook(dest: Path, workbook) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=dest.parent, delete=False) as handle:
        temp_path = Path(handle.name)
    workbook.save(temp_path)
    temp_path.replace(dest)


def _atomic_save_pdf(dest: Path, doc) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(suffix=".pdf", dir=dest.parent, delete=False) as handle:
        temp_path = Path(handle.name)
    doc.save(temp_path, garbage=4, deflate=True)
    temp_path.replace(dest)


def _reject_unsafe_pdf(doc) -> None:
    if doc.is_encrypted or doc.needs_pass:
        raise ValueError("unsupported encrypted PDF")
    if hasattr(doc, "embfile_count") and doc.embfile_count():
        raise ValueError("unsupported PDF with embedded files")
