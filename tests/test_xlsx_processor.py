import os
import subprocess
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
BIN = ROOT / "bin" / "bleach"


class XlsxProcessorTests(unittest.TestCase):
    def run_bleach(self, *args: str, bleach_home: Path) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            [str(BIN), *args],
            cwd=ROOT,
            env={**os.environ, "BLEACH_HOME": str(bleach_home)},
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=False,
        )

    def test_redact_xlsx_visible_and_hidden_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "tax.xlsx"
            out = root / "out"
            workbook = Workbook()
            workbook.active["A1"] = "SSN 123-45-6789"
            hidden = workbook.create_sheet("hidden")
            hidden.sheet_state = "hidden"
            hidden["B2"] = "email example@comcast.net"
            workbook.save(source)

            result = self.run_bleach(
                "redact",
                "--profile",
                "ai-share",
                str(source),
                "--output-dir",
                str(out),
                bleach_home=root / "state",
            )

            dest = out / "ai-share" / "tax.xlsx"
            redacted = load_workbook(dest)

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(redacted.active["A1"].value, "SSN [REDACTED:SSN]")
        self.assertEqual(redacted["hidden"]["B2"].value, "email [REDACTED:email]")

    def test_verify_xlsx_catches_residual_pii(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "tax.xlsx"
            workbook = Workbook()
            workbook.active["A1"] = "SSN 123-45-6789"
            workbook.save(source)

            result = self.run_bleach(
                "verify",
                "--profile",
                "ai-share",
                str(source),
                bleach_home=root / "state",
            )

        self.assertEqual(result.returncode, 1)
        self.assertIn("residual sensitive data", result.stdout)

    def test_xls_and_xlsm_are_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            xls = root / "old.xls"
            xlsm = root / "macro.xlsm"
            out = root / "out"
            xls.write_bytes(b"fake")
            xlsm.write_bytes(b"fake")

            result = self.run_bleach(
                "redact",
                "--profile",
                "ai-share",
                str(xls),
                str(xlsm),
                "--output-dir",
                str(out),
                bleach_home=root / "state",
            )

        self.assertEqual(result.returncode, 1)
        self.assertIn("unsupported file type", result.stdout)
        self.assertFalse((out / "ai-share" / "old.xls").exists())
        self.assertFalse((out / "ai-share" / "macro.xlsm").exists())


if __name__ == "__main__":
    unittest.main()
